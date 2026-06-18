#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel/Markdown报告生成工具
支持数据可视化、Excel表格生成、Markdown文档输出
"""

import os
import json
from typing import Dict, Any, List
from datetime import datetime
from core import BaseTool


class ReportGeneratorTool(BaseTool):
    """
    报告生成工具
    支持Excel和Markdown格式的报告生成
    """
    
    name = "report_generator"
    description = "生成Excel和Markdown格式的报告"
    parameters = {
        "content": {
            "type": "string",
            "description": "报告内容或数据",
            "required": True
        },
        "format": {
            "type": "string",
            "description": "报告格式 (excel/markdown/both)",
            "required": False
        },
        "title": {
            "type": "string",
            "description": "报告标题",
            "required": False
        },
        "output_path": {
            "type": "string",
            "description": "输出文件路径",
            "required": False
        }
    }
    
    def __init__(self, output_dir: str = "reports"):
        """
        初始化报告生成器
        :param output_dir: 报告输出目录
        """
        super().__init__()
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def _generate_markdown_report(self, content: str, title: str = "报告") -> str:
        """
        生成Markdown格式报告
        :param content: 报告内容
        :param title: 报告标题
        :return: Markdown文本
        """
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        markdown = f"""# {title}

**生成时间**: {timestamp}

---

{content}

---

*本报告由轻量化企业多工具协同自主规划办公智能Agent自动生成*
"""
        return markdown
    
    def _generate_excel_report(self, content: str, title: str = "报告") -> str:
        """
        生成Excel格式报告
        :param content: 报告内容
        :param title: 报告标题
        :return: Excel文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return "错误: 需要安装openpyxl库 (pip install openpyxl)"
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "报告"
        
        # 设置标题样式
        title_font = Font(size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入标题
        ws['A1'] = title
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = title_alignment
        ws.merge_cells('A1:D1')
        
        # 写入生成时间
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A2'] = f"生成时间: {timestamp}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:D2')
        
        # 解析内容并写入
        row = 4
        
        # 尝试解析JSON格式的数据
        try:
            data = json.loads(content)
            
            if isinstance(data, list):
                # 表格数据
                if data and isinstance(data[0], dict):
                    # 写入表头
                    headers = list(data[0].keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    
                    row += 1
                    
                    # 写入数据行
                    for item in data:
                        for col, header in enumerate(headers, 1):
                            value = item.get(header, "")
                            ws.cell(row=row, column=col, value=str(value))
                        row += 1
            
            elif isinstance(data, dict):
                # 键值对数据
                for key, value in data.items():
                    ws.cell(row=row, column=1, value=str(key))
                    ws.cell(row=row, column=2, value=str(value))
                    row += 1
            
        except json.JSONDecodeError:
            # 普通文本内容
            lines = content.split('\n')
            for line in lines:
                if line.strip():
                    ws.cell(row=row, column=1, value=line)
                    ws.merge_cells(f'A{row}:D{row}')
                    row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        
        # 生成文件名
        filename = f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # 保存文件
        wb.save(filepath)
        
        return filepath
    
    def _generate_data_table(self, data: List[Dict[str, Any]]) -> str:
        """
        将数据转换为Markdown表格
        :param data: 数据列表
        :return: Markdown表格字符串
        """
        if not data:
            return "无数据"
        
        # 获取表头
        headers = list(data[0].keys())
        
        # 构建表格
        table = "| " + " | ".join(headers) + " |\n"
        table += "| " + " | ".join(["---"] * len(headers)) + " |\n"
        
        # 添加数据行
        for row in data:
            values = [str(row.get(h, "")) for h in headers]
            table += "| " + " | ".join(values) + " |\n"
        
        return table
    
    def _parse_content(self, content: str) -> Dict[str, Any]:
        """
        解析内容，提取结构化数据
        :param content: 原始内容
        :return: 解析后的数据字典
        """
        parsed = {
            "text": content,
            "data": None,
            "tables": []
        }
        
        # 尝试解析JSON
        try:
            data = json.loads(content)
            parsed["data"] = data
        except json.JSONDecodeError:
            pass
        
        # 查找表格数据（简单的Markdown表格检测）
        lines = content.split('\n')
        current_table = []
        in_table = False
        
        for line in lines:
            if line.strip().startswith('|'):
                if not in_table:
                    in_table = True
                current_table.append(line)
            elif in_table:
                if current_table:
                    parsed["tables"].append("\n".join(current_table))
                current_table = []
                in_table = False
        
        if current_table:
            parsed["tables"].append("\n".join(current_table))
        
        return parsed
    
    def _format_result(self, markdown_content: str, excel_path: str = None) -> str:
        """
        格式化生成结果
        :param markdown_content: Markdown内容
        :param excel_path: Excel文件路径
        :return: 格式化结果文本
        """
        result = "✅ 报告生成成功\n\n"
        
        result += "📄 Markdown报告:\n"
        result += "-" * 40 + "\n"
        result += markdown_content[:500]
        if len(markdown_content) > 500:
            result += "\n... (内容已截断)"
        result += "\n" + "-" * 40 + "\n\n"
        
        if excel_path:
            result += f"📊 Excel报告: {excel_path}\n"
        
        return result
    
    def execute(self, **kwargs) -> str:
        """
        生成报告
        :param kwargs: 包含content, format, title, output_path等参数
        :return: 生成结果文本
        """
        content = kwargs.get("content", "")
        format_type = kwargs.get("format", "both").lower()
        title = kwargs.get("title", "报告")
        output_path = kwargs.get("output_path", "")
        
        if not content:
            return "错误: 报告内容不能为空"
        
        print(f"[报告生成] 格式: {format_type}")
        print(f"[报告生成] 标题: {title}")
        
        # 解析内容
        parsed = self._parse_content(content)
        
        # 生成Markdown报告
        markdown_content = self._generate_markdown_report(content, title)
        
        # 生成Excel报告
        excel_path = None
        if format_type in ["excel", "both"]:
            excel_path = self._generate_excel_report(content, title)
            if excel_path.startswith("错误"):
                print(f"[报告生成] Excel生成失败: {excel_path}")
            else:
                print(f"[报告生成] Excel已保存: {excel_path}")
        
        # 保存Markdown报告
        if format_type in ["markdown", "both"]:
            filename = f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
            filepath = output_path or os.path.join(self.output_dir, filename)
            
            try:
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(markdown_content)
                print(f"[报告生成] Markdown已保存: {filepath}")
            except Exception as e:
                print(f"[报告生成] Markdown保存失败: {str(e)}")
        
        # 格式化结果
        result = self._format_result(markdown_content, excel_path)
        
        print("[报告生成] 完成")
        
        return result


def create_report_generator_tool(output_dir: str = "reports") -> ReportGeneratorTool:
    """
    工厂函数：创建报告生成器工具实例
    :param output_dir: 报告输出目录
    :return: ReportGeneratorTool实例
    """
    return ReportGeneratorTool(output_dir)